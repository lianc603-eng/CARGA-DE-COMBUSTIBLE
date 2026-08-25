import streamlit as st
import pandas as pd
from datetime import datetime, date, time
import pytz
import requests
import json
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

st.set_page_config(page_title="Control de Combustible", layout="wide", page_icon="⛽")

# ==========================================================
# CONFIGURACIONES GENERALES Y HORARIO
# ==========================================================
PRESUPUESTO_GLOBAL = 3800.00
BOLSA_COMODIN_TOTAL = 200.00
HORA_LIMITE = time(15, 10)  # ⏰ 3:10 PM
ZONA_HORARIA = pytz.timezone("America/Merida")

CONFIG_FILE = "config_sistema.json"
WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbzOjgha2Zjyog01t6LmA_R--EB4Ecqv2ifO_i2YJbLRLbXGShbu5uzFVi85FUTGplM8/exec"

PRESUPUESTO_BASE_POR_SOLICITANTE = {
    "COB CHAVEZ NARCISO DEL JESUS": 200.00,
    "PEREZ MAZIN CARLOS EDUARDO": 200.00,
    "DE LA CRUZ PEREZ WILLIAN ARLEY": 200.00,
    "NOEL CHAN": 850.00,
    "LIAN": 150.00,
    "QUEVEDO": 1500.00,
    "RENAN/HELDER": 500.00,
}

OPERADORES_POR_SOLICITANTE = {
    "COB CHAVEZ NARCISO DEL JESUS": ["JESUS COB"],
    "PEREZ MAZIN CARLOS EDUARDO": ["EDUARDO PEREZ"],
    "DE LA CRUZ PEREZ WILLIAN ARLEY": ["WILLIAN PEREZ"],
    "NOEL CHAN": ["AXEL SARAVIA", "NOEL CHAN", "ROMAN DZUL", "ROGER DUARTE", "LUIS CAHUICH"],
    "LIAN": ["FRANCISCO ALONZO"],
    "QUEVEDO": ["FARID PAVON", "WALDEMAR SAGUNDO", "JORGE MELIK"],
    "RENAN/HELDER": ["HELDER PACHECO", "RENAN CETINA"],
}

TXT_DESARROLLO_URBANO = "LLEVAR A CABO ACTIVIDADES DE INSPECCIONES, VERIFICACIONES Y SUPERVICIONES DE OBRAS Y OBSTRUCCIONES A LA VIA PÚBLICA CORRESPONDIENTES A LA SUBDIRECCION DE DESARROLLO URBANO"
TXT_MEDIO_AMBIENTE = "PARA LLEVAR A CABO INSPECCIONES A CARGO DE LA SUBDIRECCION DE MEDIO AMBIENTE, COMO LO SON ATENDER REPORTES POR TIRADERO DE AGUAS JABONOSAS, MALTRATO ANIMAL Y CONTAMINACION AUDITIVA, ASI COMO DIVERSOS TIPOS DE CONTAMINACION"
TXT_RAM_AMBIENTAL = "PARA LLEVAR A CABO ACTIVIDADES DE ESTERILIZACIONES DE PERROS Y GATOS, RECOLECCION DE MERMA DE FRUTAS Y VERDURAS EN SUPERMERCADOS Y REFORESTACIONES"

MAPEO_SOLICITANTES = {
    12: {"solicita": "COB CHAVEZ NARCISO DEL JESUS", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "85GWU7", "actividad": TXT_DESARROLLO_URBANO},
    13: {"solicita": "PEREZ MAZIN CARLOS EDUARDO", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "86GWU7", "actividad": TXT_DESARROLLO_URBANO},
    14: {"solicita": "DE LA CRUZ PEREZ WILLIAN ARLEY", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "86GWU8", "actividad": TXT_DESARROLLO_URBANO},
    15: {"solicita": "COB CHAVEZ NARCISO DEL JESUS", "vehiculo": "MOTOCICLETA HONDA", "placa": "88GWU7", "actividad": TXT_DESARROLLO_URBANO},
    16: {"solicita": "NOEL CHAN", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "88GWU8", "actividad": TXT_MEDIO_AMBIENTE},
    17: {"solicita": "NOEL CHAN", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "89GWU7", "actividad": TXT_MEDIO_AMBIENTE},
    18: {"solicita": "NOEL CHAN", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "89GWU8", "actividad": TXT_MEDIO_AMBIENTE},
    19: {"solicita": "NOEL CHAN", "vehiculo": "MOTOCICLETA HONDA", "placa": "90GWU7", "actividad": TXT_MEDIO_AMBIENTE},
    20: {"solicita": "NOEL CHAN", "vehiculo": "MOTOCICLETA HONDA", "placa": "90GWU8", "actividad": TXT_MEDIO_AMBIENTE},
    21: {"solicita": "LIAN", "vehiculo": "MOTOCICLETA SUZUKI", "placa": "91GWU7", "actividad": TXT_MEDIO_AMBIENTE},
    22: {"solicita": "QUEVEDO", "vehiculo": "CAMIONETA RAM", "placa": "CN2633B", "actividad": TXT_RAM_AMBIENTAL},
    23: {"solicita": "RENAN/HELDER", "vehiculo": "AUTOMOVIL JETTA", "placa": "DFT565C", "actividad": TXT_DESARROLLO_URBANO},
}

USUARIOS_PASSWORD = {
    "LIAN": "admin123",
    "NOEL CHAN": "inspeccion2026",
    "QUEVEDO": "ambiental2026",
    "RENAN/HELDER": "urbano2026",
    "COB CHAVEZ NARCISO DEL JESUS": "notif123",
    "PEREZ MAZIN CARLOS EDUARDO": "notif123",
    "DE LA CRUZ PEREZ WILLIAN ARLEY": "notif123",
}

# --- PERSISTENCIA LOCAL ---
def leer_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "desbloqueo_horario": False,
        "asignacion_comodin": {},
        "cesion_lian": {},
        "dia_activo": "Lunes"
    }

def guardar_config(cfg):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(cfg, f)
    except Exception:
        pass

def calcular_presupuesto_efectivo():
    cfg = leer_config()
    asig_comodin = cfg.get("asignacion_comodin", {})
    cesion_lian = cfg.get("cesion_lian", {})
    
    presupuestos = PRESUPUESTO_BASE_POR_SOLICITANTE.copy()
    
    for solicitante, extra in asig_comodin.items():
        if solicitante in presupuestos:
            presupuestos[solicitante] += float(extra)
            
    total_cedido_lian = 0.0
    for solicitante, monto in cesion_lian.items():
        if solicitante in presupuestos and solicitante != "LIAN":
            presupuestos[solicitante] += float(monto)
            total_cedido_lian += float(monto)
            
    presupuestos["LIAN"] = max(0.0, PRESUPUESTO_BASE_POR_SOLICITANTE["LIAN"] - total_cedido_lian)
    return presupuestos

def limpiar_texto_operador(val):
    if val is None or pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() in ["none", "null", "nan", ""] else s

# --- GESTIÓN DE DATOS ---
def obtener_datos_sheets(forzar=False):
    if "df_datos_persistentes" in st.session_state and not forzar:
        return st.session_state.df_datos_persistentes.copy()

    filas_dict = {}
    try:
        res = requests.get(WEBHOOK_URL, timeout=8, allow_redirects=True)
        if res.status_code == 200:
            datos_raw = res.json().get("data", [])
            for item in datos_raw:
                r = int(item.get("row", 0))
                if r in MAPEO_SOLICITANTES:
                    op_l = limpiar_texto_operador(item.get("encargado_lunes") or item.get("encargado_solicitado"))
                    imp_l = item.get("importe_lunes") if item.get("importe_lunes") is not None else item.get("importe_solicitado", 0.0)
                    real_l = item.get("real_lunes", 0.0)
                    
                    op_j = limpiar_texto_operador(item.get("encargado_jueves") or item.get("encargado_real"))
                    imp_j = item.get("importe_jueves", 0.0)
                    real_j = item.get("real_jueves", 0.0)

                    filas_dict[r] = {
                        "row": r,
                        "Solicitante": MAPEO_SOLICITANTES[r]["solicita"],
                        "Vehículo": MAPEO_SOLICITANTES[r]["vehiculo"],
                        "Placa": MAPEO_SOLICITANTES[r]["placa"],
                        "Actividad": MAPEO_SOLICITANTES[r]["actividad"],
                        "Operador_Lunes": op_l,
                        "Importe_Lunes": float(imp_l) if imp_l else 0.0,
                        "Real_Lunes": float(real_l) if real_l else 0.0,
                        "Operador_Jueves": op_j,
                        "Importe_Jueves": float(imp_j) if imp_j else 0.0,
                        "Real_Jueves": float(real_j) if real_j else 0.0
                    }
    except Exception:
        pass

    filas_finales = []
    for r in sorted(MAPEO_SOLICITANTES.keys()):
        if r in filas_dict:
            filas_finales.append(filas_dict[r])
        else:
            info = MAPEO_SOLICITANTES[r]
            filas_finales.append({
                "row": r,
                "Solicitante": info["solicita"],
                "Vehículo": info["vehiculo"],
                "Placa": info["placa"],
                "Actividad": info["actividad"],
                "Operador_Lunes": "",
                "Importe_Lunes": 0.0,
                "Real_Lunes": 0.0,
                "Operador_Jueves": "",
                "Importe_Jueves": 0.0,
                "Real_Jueves": 0.0
            })
            
    df_res = pd.DataFrame(filas_finales)
    st.session_state.df_datos_persistentes = df_res.copy()
    return df_res

def enviar_datos_sheets(registros_a_enviar, turno="lunes", f_elab=None, f_prog=None, historico_obj=None):
    payload = {"tipo": turno, "turno": turno, "registros": []}
    if f_elab:
        payload["fecha_elaboro"] = f_elab.strftime("%d/%m/%Y")
    if f_prog:
        payload["fecha_prog"] = f_prog.strftime("%d/%m/%Y")
    if historico_obj:
        payload["historico"] = historico_obj
        
    for _, fila in registros_a_enviar.iterrows():
        if turno == "lunes":
            enc = fila["Operador_Lunes"]
            imp = fila["Importe_Lunes"]
        elif turno == "real_lunes":
            enc = fila["Operador_Lunes"]
            imp = fila["Real_Lunes"]
        elif turno == "jueves":
            enc = fila["Operador_Jueves"]
            imp = fila["Importe_Jueves"]
        elif turno == "real_jueves":
            enc = fila["Operador_Jueves"]
            imp = fila["Real_Jueves"]
        else:
            enc = ""
            imp = 0.0

        payload["registros"].append({
            "row": int(fila["row"]),
            "encargado": limpiar_texto_operador(enc),
            "importe": float(imp) if pd.notna(imp) else 0.0
        })
        
    if "df_datos_persistentes" in st.session_state:
        df_mem = st.session_state.df_datos_persistentes
        for _, r_env in registros_a_enviar.iterrows():
            mask = df_mem["row"] == r_env["row"]
            if turno == "lunes":
                df_mem.loc[mask, "Operador_Lunes"] = limpiar_texto_operador(r_env["Operador_Lunes"])
                df_mem.loc[mask, "Importe_Lunes"] = r_env["Importe_Lunes"]
            elif turno == "real_lunes":
                df_mem.loc[mask, "Real_Lunes"] = r_env["Real_Lunes"]
            elif turno == "jueves":
                df_mem.loc[mask, "Operador_Jueves"] = limpiar_texto_operador(r_env["Operador_Jueves"])
                df_mem.loc[mask, "Importe_Jueves"] = r_env["Importe_Jueves"]
            elif turno == "real_jueves":
                df_mem.loc[mask, "Real_Jueves"] = r_env["Real_Jueves"]
        st.session_state.df_datos_persistentes = df_mem

    try:
        res = requests.post(WEBHOOK_URL, json=payload, timeout=12, allow_redirects=True)
        if res.status_code == 200:
            return res.json().get("status") == "success"
        return False
    except Exception:
        return False

# ==========================================
# GENERADORES DE ARCHIVOS OFICIALES (LUNES / JUEVES)
# ==========================================
def generar_excel_oficial_formato(df_datos, dia_reporte, f_elab, f_prog):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = "carga"
    ws.views.sheetView[0].showGridLines = True
    
    fuente_titulo = Font(name="Calibri", size=10, bold=True)
    fuente_sub = Font(name="Calibri", size=9, bold=True)
    fuente_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    fuente_bold = Font(name="Calibri", size=9, bold=True)
    fuente_datos = Font(name="Calibri", size=8)
    
    fill_header_azul = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border_cuadricula = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )

    ws["D2"] = "H. AYUNTAMIENTO DE CAMPECHE"
    ws["D2"].font = fuente_titulo
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["B4"] = "Unidad:"
    ws["B4"].font = fuente_sub
    ws["C4"] = "DIRECCION DE DESARROLLO URBANO Y MEDIO AMBIENTE"
    ws["C4"].font = fuente_sub
    
    ws["B5"] = "Subdireccion:"
    ws["B5"].font = fuente_sub
    
    ws["H8"] = "Elaboro:"
    ws["H8"].font = fuente_sub
    ws["H8"].alignment = Alignment(horizontal="right")
    ws["I8"] = f_elab.strftime("%d/%m/%Y")
    ws["I8"].font = fuente_sub
    
    ws["H9"] = "Programacion para el dia:"
    ws["H9"].font = fuente_sub
    ws["H9"].alignment = Alignment(horizontal="right")
    ws["I9"] = f_prog.strftime("%d/%m/%Y")
    ws["I9"].font = fuente_sub

    headers_oficiales = [
        (2, "NOMBRE DEL\nENCARGADO"), (3, "VEHÍCULO"), (4, "PLACA"),
        (5, "OFICIAL /\nCOMODATO"), (6, "LITROS"), (7, "IMPORTE"),
        (8, "MAGNA / DIESEL"), (9, "ACTIVIDAD")
    ]
    
    for col_num, h_text in headers_oficiales:
        cell = ws.cell(row=11, column=col_num, value=h_text)
        cell.font = fuente_header
        cell.fill = fill_header_azul
        cell.border = border_cuadricula
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[11].height = 28

    col_op = "Operador_Lunes" if dia_reporte == "Lunes" else "Operador_Jueves"
    col_imp = "Importe_Lunes" if dia_reporte == "Lunes" else "Importe_Jueves"

    for r_num in sorted(MAPEO_SOLICITANTES.keys()):
        row_info = df_datos[df_datos["row"] == r_num]
        
        if not row_info.empty:
            item = row_info.iloc[0]
            op_nombre = limpiar_texto_operador(item[col_op])
            veh = str(item["Vehículo"]).strip()
            plc = str(item["Placa"]).strip()
            imp_val = float(item[col_imp])
            act_text = str(item["Actividad"]).strip()
        else:
            op_nombre = ""
            veh = MAPEO_SOLICITANTES[r_num]["vehiculo"]
            plc = MAPEO_SOLICITANTES[r_num]["placa"]
            imp_val = 0.0
            act_text = MAPEO_SOLICITANTES[r_num]["actividad"]

        ws.cell(row=r_num, column=2, value=op_nombre).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r_num, column=3, value=veh).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r_num, column=4, value=plc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_num, column=5, value="OFICIAL").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_num, column=6, value="").alignment = Alignment(horizontal="center", vertical="center")
        
        c_imp = ws.cell(row=r_num, column=7, value=imp_val)
        c_imp.number_format = '$#,##0.00'
        c_imp.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=r_num, column=8, value="MAGNA").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_num, column=9, value=act_text).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col_i in range(2, 10):
            c_est = ws.cell(row=r_num, column=col_i)
            c_est.font = fuente_datos
            c_est.border = border_cuadricula
            
        ws.row_dimensions[r_num].height = 24
        if imp_val == 0.0:
            ws.row_dimensions[r_num].hidden = True

    c_lbl_tot = ws.cell(row=24, column=6, value="TOTAL")
    c_lbl_tot.font = fuente_bold
    c_lbl_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_lbl_tot.border = border_cuadricula
    
    c_val_tot = ws.cell(row=24, column=7, value="=SUM(G12:G23)")
    c_val_tot.font = fuente_bold
    c_val_tot.number_format = '$#,##0.00'
    c_val_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_val_tot.border = border_cuadricula
    
    for c_rest in [2, 3, 4, 5, 8, 9]:
        ws.cell(row=24, column=c_rest).border = border_cuadricula

    anchos_cols = {'A': 3, 'B': 24, 'C': 18, 'D': 12, 'E': 14, 'F': 10, 'G': 14, 'H': 14, 'I': 52}
    for col_letra, ancho in anchos_cols.items():
        ws.column_dimensions[col_letra].width = ancho

    wb.save(output)
    return output.getvalue()

def generar_pdf_oficial(df_cargas, dia_reporte, f_elab, f_prog):
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 6, 'H. AYUNTAMIENTO DE CAMPECHE', 0, 1, 'C')
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 5, f'DIRECCION DE DESARROLLO URBANO Y MEDIO AMBIENTE - OFICIO DE CARGA {dia_reporte.upper()}', 0, 1, 'C')
    pdf.ln(3)
    
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(140, 5, 'Unidad: DIRECCION DE DESARROLLO URBANO Y MEDIO AMBIENTE', 0, 0, 'L')
    pdf.cell(130, 5, f'Elaboro: {f_elab.strftime("%d/%m/%Y")}', 0, 1, 'R')
    pdf.cell(140, 5, '', 0, 0, 'L')
    pdf.cell(130, 5, f'Programacion para el dia: {f_prog.strftime("%d/%m/%Y")}', 0, 1, 'R')
    pdf.ln(4)
    
    col_widths = (45, 32, 18, 18, 22, 16, 117)
    col_op = "Operador_Lunes" if dia_reporte == "Lunes" else "Operador_Jueves"
    col_imp = "Importe_Lunes" if dia_reporte == "Lunes" else "Importe_Jueves"
    
    with pdf.table(
        col_widths=col_widths, 
        text_align=("LEFT", "LEFT", "CENTER", "CENTER", "RIGHT", "CENTER", "LEFT"),
        line_height=4.5
    ) as table:
        pdf.set_font('Helvetica', 'B', 8)
        header_row = table.row()
        for h in ['NOMBRE DEL ENCARGADO', 'VEHICULO', 'PLACA', 'REGIMEN', 'IMPORTE', 'TIPO', 'ACTIVIDAD']:
            header_row.cell(h)
            
        pdf.set_font('Helvetica', '', 7.5)
        total = 0.0
        for _, row in df_cargas.iterrows():
            data_row = table.row()
            data_row.cell(limpiar_texto_operador(row[col_op]))
            data_row.cell(str(row["Vehículo"]).strip())
            data_row.cell(str(row["Placa"]).strip())
            data_row.cell('OFICIAL')
            data_row.cell(f"${float(row[col_imp]):,.2f}")
            data_row.cell('MAGNA')
            data_row.cell(str(row["Actividad"]).strip())
            total += float(row[col_imp])
            
        pdf.set_font('Helvetica', 'B', 8)
        tot_row = table.row()
        tot_row.cell(f'TOTAL AUTORIZADO {dia_reporte.upper()}:', colspan=4, align="RIGHT")
        tot_row.cell(f"${total:,.2f}", align="RIGHT")
        tot_row.cell('', colspan=2)
        
    return bytes(pdf.output())

# ==========================================
# 1. INICIO DE SESIÓN
# ==========================================
if "usuario_logueado" not in st.session_state:
    st.session_state.usuario_logueado = None
if "vista_simulada" not in st.session_state:
    st.session_state.vista_simulada = None

if st.session_state.usuario_logueado is None:
    st.title("⛽ Solicitud de Combustible")
    st.caption("Dirección de Desarrollo Urbano y Medio Ambiente")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("form_login"):
            st.subheader("🔐 Iniciar Sesión")
            usr = st.selectbox("Selecciona tu Usuario", list(USUARIOS_PASSWORD.keys()))
            pwd = st.text_input("Contraseña", type="password")
            if st.form_submit_button("Entrar", use_container_width=True):
                if pwd == USUARIOS_PASSWORD.get(usr):
                    st.session_state.usuario_logueado = usr
                    st.session_state.vista_simulada = None
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta.")
    st.stop()

# ==========================================
# 2. ENCABEZADO Y CONFIGURACIÓN
# ==========================================
usuario_real = st.session_state.usuario_logueado
es_admin_real = (usuario_real == "LIAN")
usuario_efectivo = st.session_state.vista_simulada if (es_admin_real and st.session_state.vista_simulada) else usuario_real
es_admin = (usuario_efectivo == "LIAN")

ahora_local = datetime.now(ZONA_HORARIA)
hora_actual = ahora_local.time()

cfg_actual = leer_config()
desbloqueo_activo = cfg_actual.get("desbloqueo_horario", False)
sistema_bloqueado = (hora_actual >= HORA_LIMITE) and not es_admin_real and not desbloqueo_activo

presupuestos_actuales = calcular_presupuesto_efectivo()

c1, c2, c3 = st.columns([2.5, 1.2, 0.8])
with c1:
    st.title("⛽ Control de Combustible")
    if es_admin_real and st.session_state.vista_simulada:
        st.warning(f"🧪 **MODO DE PRUEBA ACTIVO**: Simulando vista como **{usuario_efectivo}**")
    else:
        st.markdown("👑 **ADMINISTRADOR GENERAL**" if es_admin else f"👤 Solicitante: **{usuario_efectivo}**")
        
    estado_horario = "🟢 Horario Abierto" if not sistema_bloqueado else "🔴 Horario Cerrado (Límite 3:10 PM)"
    if desbloqueo_activo:
        estado_horario += " (⚡ Desbloqueo temporal activo)"
    st.caption(f"🕒 {ahora_local.strftime('%I:%M %p')} | {estado_horario}")

with c2:
    dia_guardado = cfg_actual.get("dia_activo", "Lunes")
    if es_admin:
        dia_activo = st.radio("📅 Turno de Captura para Usuarios:", ["Lunes", "Jueves"], horizontal=True, index=0 if dia_guardado == "Lunes" else 1)
        if dia_activo != dia_guardado:
            cfg_actual["dia_activo"] = dia_activo
            guardar_config(cfg_actual)
            st.rerun()
    else:
        dia_activo = dia_guardado
        st.info(f"📅 Solicitud Abierta para: **{dia_activo}**")

with c3:
    st.write("")
    if st.button("🔄 Sincronizar", use_container_width=True):
        obtener_datos_sheets(forzar=True)
        st.toast("Datos sincronizados con Google Sheets.", icon="✅")
        st.rerun()
    if es_admin_real and st.session_state.vista_simulada:
        if st.button("⬅️ Volver", use_container_width=True):
            st.session_state.vista_simulada = None
            st.rerun()
    else:
        if st.button("🚪 Salir", use_container_width=True):
            st.session_state.usuario_logueado = None
            st.session_state.vista_simulada = None
            st.rerun()

df_actual = obtener_datos_sheets()

# ==========================================
# 3. VISTA SOLICITANTE (LUNES / JUEVES CONTROLADO)
# ==========================================
if not es_admin:
    presupuesto_semanal_total = presupuestos_actuales.get(usuario_efectivo, 0.00)
    df_solicitante = df_actual[df_actual["Solicitante"] == usuario_efectivo].copy()
    
    # Si es Lunes: Saldo disponible es el total semanal
    # Si es Jueves: Saldo disponible = Total Semanal - Real Lunes (o Importe Lunes si aún no auditan)
    total_lunes_gastado = df_solicitante["Real_Lunes"].sum()
    if total_lunes_gastado == 0.0:
        total_lunes_gastado = df_solicitante["Importe_Lunes"].sum()
        
    if dia_activo == "Lunes":
        saldo_disponible_hoy = presupuesto_semanal_total
        gasto_previo_info = 0.0
    else:
        saldo_disponible_hoy = max(0.0, presupuesto_semanal_total - total_lunes_gastado)
        gasto_previo_info = total_lunes_gastado

    lista_operadores_autorizados = OPERADORES_POR_SOLICITANTE.get(usuario_efectivo, [])
    
    if sistema_bloqueado:
        st.error("🔒 **SISTEMA CERRADO POR HORARIO (3:10 PM)**. La captura ha finalizado.")
    else:
        if dia_activo == "Jueves":
            st.info(f"💡 Se registraron **${gasto_previo_info:,.2f}** cargados el Lunes. Tu saldo restante disponible para este Jueves es **${saldo_disponible_hoy:,.2f}**.")
        st.caption(f"📱 Selecciona los conductores e importes para la carga del **{dia_activo}**:")
    
    with st.form("form_solicitante_movil"):
        nuevos_valores = []
        col_op_target = "Operador_Lunes" if dia_activo == "Lunes" else "Operador_Jueves"
        col_imp_target = "Importe_Lunes" if dia_activo == "Lunes" else "Importe_Jueves"
        
        for idx, row in df_solicitante.iterrows():
            with st.container(border=True):
                st.markdown(f"🛵 **{row['Vehículo']}** &nbsp;|&nbsp; Placa: **`{row['Placa']}`**")
                st.caption(f"📋 **Actividad:** {row['Actividad']}")
                
                c_op, c_imp = st.columns([1.5, 1])
                opciones_operadores = [""] + lista_operadores_autorizados
                val_actual = limpiar_texto_operador(row[col_op_target])
                
                if val_actual and val_actual not in opciones_operadores:
                    opciones_operadores.append(val_actual)
                    
                idx_sel = opciones_operadores.index(val_actual) if val_actual in opciones_operadores else 0
                
                with c_op:
                    val_encargado = st.selectbox(
                        "Operador / Conductor",
                        options=opciones_operadores,
                        index=idx_sel,
                        key=f"op_{dia_activo}_{row['row']}",
                        disabled=sistema_bloqueado
                    )
                    
                with c_imp:
                    val_importe = st.number_input(
                        f"Monto {dia_activo} ($)",
                        value=float(row[col_imp_target]),
                        step=50.0,
                        min_value=0.0,
                        key=f"imp_{dia_activo}_{row['row']}",
                        disabled=sistema_bloqueado,
                        format="%.2f"
                    )
                
                nuevos_valores.append({
                    "row": row["row"],
                    "Solicitante": row["Solicitante"],
                    "Vehículo": row["Vehículo"],
                    "Placa": row["Placa"],
                    "Actividad": row["Actividad"],
                    "Operador_Lunes": val_encargado if dia_activo == "Lunes" else row["Operador_Lunes"],
                    "Importe_Lunes": val_importe if dia_activo == "Lunes" else row["Importe_Lunes"],
                    "Real_Lunes": row["Real_Lunes"],
                    "Operador_Jueves": val_encargado if dia_activo == "Jueves" else row["Operador_Jueves"],
                    "Importe_Jueves": val_importe if dia_activo == "Jueves" else row["Importe_Jueves"],
                    "Real_Jueves": row["Real_Jueves"]
                })
        
        df_edit_movil = pd.DataFrame(nuevos_valores)
        total_capturado_hoy = df_edit_movil[col_imp_target].sum()
        saldo_restante_hoy = saldo_disponible_hoy - total_capturado_hoy
        
        st.divider()
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Presupuesto Semanal Total", f"${presupuesto_semanal_total:,.2f}")
        m2.metric("Gasto Previo Lunes", f"${gasto_previo_info:,.2f}")
        m3.metric(f"Capturado {dia_activo}", f"${total_capturado_hoy:,.2f}")
        m4.metric("Saldo Restante", f"${saldo_restante_hoy:,.2f}", delta_color="normal" if saldo_restante_hoy >= 0 else "off")
        
        if total_capturado_hoy > saldo_disponible_hoy:
            st.error(f"⚠️ Excedes el saldo disponible para este {dia_activo} por **${abs(saldo_restante_hoy):,.2f} MXN**.")
            
        btn_guardar = st.form_submit_button(f"💾 Guardar Solicitud de {dia_activo}", type="primary", use_container_width=True, disabled=sistema_bloqueado)
        
        if btn_guardar:
            if total_capturado_hoy > saldo_disponible_hoy:
                st.error("No puedes guardar si excedes el saldo disponible.")
            else:
                with st.spinner(f"Guardando solicitud del {dia_activo}..."):
                    exito = enviar_datos_sheets(df_edit_movil, turno=dia_activo.lower())
                    if exito:
                        st.success(f"✅ ¡Solicitud del {dia_activo} guardada con éxito en Google Sheets!")
                    else:
                        st.error("Error al comunicarse con Google Sheets.")

# ==========================================
# 4. VISTA ADMINISTRADOR (LIAN)
# ==========================================
else:
    st.subheader("⚙️ Panel de Consolidación, Edición y Descarga Oficial")
    
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        f_elab = st.date_input("Fecha de Elaboración", value=date.today())
    with col_f2:
        f_prog = st.date_input("Programación para el día", value=date.today())

    tab_saldos, tab_lunes, tab_jueves, tab_mi_carga, tab_auditoria, tab_mantenimiento = st.tabs([
        "📊 Monitoreo y Asignación de Presupuestos",
        "📄 Carga LUNES (Formato y Edición)",
        "📄 Carga JUEVES (Formato y Edición)",
        "🛵 Mi Carga (LIAN / Alonzo)",
        "✅ Cierre Semanal y Auditoría",
        "🛠️ Modo Pruebas y Cierre de Viernes"
    ])

    # 1. MONITOREO Y TRANSFERENCIA DE PRESUPUESTOS
    with tab_saldos:
        st.markdown("##### 💵 Balance Semanal de Presupuestos en Tiempo Real")
        
        tot_lunes_sol = df_actual["Importe_Lunes"].sum()
        tot_jueves_sol = df_actual["Importe_Jueves"].sum()
        tot_global_semana = tot_lunes_sol + tot_jueves_sol
        saldo_global_semana = PRESUPUESTO_GLOBAL - tot_global_semana
        
        asig_comodin_dict = cfg_actual.get("asignacion_comodin", {})
        total_comodin_usado = sum(asig_comodin_dict.values())
        comodin_disponible = max(0.0, BOLSA_COMODIN_TOTAL - total_comodin_usado)
        
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Presupuesto Semanal Global", f"${PRESUPUESTO_GLOBAL:,.2f}")
        c2.metric("Total Solicitado Lunes", f"${tot_lunes_sol:,.2f}")
        c3.metric("Total Solicitado Jueves", f"${tot_jueves_sol:,.2f}")
        c4.metric(
            "Saldo Libre Semana", 
            f"${saldo_global_semana:,.2f}", 
            delta=f"${tot_global_semana:,.2f} total pedido",
            delta_color="normal" if saldo_global_semana >= 0 else "inverse"
        )
        
        st.write("")
        st.markdown("##### 📋 Resumen Financiero por Solicitante")
        filas_saldos_area = []
        for sol, p_efectivo in presupuestos_actuales.items():
            sub_df = df_actual[df_actual["Solicitante"] == sol]
            sol_l = sub_df["Importe_Lunes"].sum()
            sol_j = sub_df["Importe_Jueves"].sum()
            sol_tot = sol_l + sol_j
            disp_monto = p_efectivo - sol_tot
            
            if disp_monto == 0:
                estatus = "✅ 100% Ejercido"
            elif disp_monto < 0:
                estatus = "⚠️ Excedido"
            elif sol_tot > 0:
                estatus = "🟢 Con Saldo"
            else:
                estatus = "⚪ Sin Carga"
                
            filas_saldos_area.append({
                "Solicitante / Área": sol,
                "Presupuesto Base": PRESUPUESTO_BASE_POR_SOLICITANTE.get(sol, 0.0),
                "Presupuesto Semanal Autorizado": p_efectivo,
                "Carga Lunes": sol_l,
                "Carga Jueves": sol_j,
                "Total Solicitado": sol_tot,
                "Saldo Disponible Restante": disp_monto,
                "Estatus": estatus
            })
            
        df_saldos_area = pd.DataFrame(filas_saldos_area)
        st.dataframe(
            df_saldos_area,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Presupuesto Base": st.column_config.NumberColumn(format="$%.2f"),
                "Presupuesto Semanal Autorizado": st.column_config.NumberColumn(format="$%.2f"),
                "Carga Lunes": st.column_config.NumberColumn(format="$%.2f"),
                "Carga Jueves": st.column_config.NumberColumn(format="$%.2f"),
                "Total Solicitado": st.column_config.NumberColumn(format="$%.2f"),
                "Saldo Disponible Restante": st.column_config.NumberColumn(format="$%.2f"),
            }
        )

        st.divider()
        st.markdown("##### 🔀 Asignar Comodín ($200) y Ceder Presupuesto de Francisco Alonzo / LIAN ($150)")
        col_trans1, col_trans2 = st.columns(2)
        
        with col_trans1:
            with st.container(border=True):
                st.markdown(f"🎁 **Bolsa Comodín (Libre: ${comodin_disponible:,.2f})**")
                areas_comodin = [u for u in PRESUPUESTO_BASE_POR_SOLICITANTE.keys() if u != "LIAN"]
                destinatario_comodin = st.selectbox("Asignar Comodín a:", areas_comodin, key="sel_comodin")
                monto_comodin_add = st.number_input("Monto Extra ($)", min_value=0.0, max_value=float(comodin_disponible), step=50.0, value=0.0, key="inp_comodin")
                
                if st.button("➕ Asignar Extra de Comodín", use_container_width=True):
                    if monto_comodin_add > 0:
                        asig_act = cfg_actual.get("asignacion_comodin", {})
                        asig_act[destinatario_comodin] = asig_act.get(destinatario_comodin, 0.0) + monto_comodin_add
                        cfg_actual["asignacion_comodin"] = asig_act
                        guardar_config(cfg_actual)
                        st.toast(f"Se asignaron ${monto_comodin_add:,.2f} a {destinatario_comodin}", icon="🎁")
                        st.rerun()

        with col_trans2:
            with st.container(border=True):
                presupuesto_lian_actual = presupuestos_actuales["LIAN"]
                st.markdown(f"🤝 **Ceder Presupuesto Francisco Alonzo / LIAN (Disponible: ${presupuesto_lian_actual:,.2f})**")
                areas_ceder = [u for u in PRESUPUESTO_BASE_POR_SOLICITANTE.keys() if u != "LIAN"]
                destinatario_ceder = st.selectbox("Ceder a:", areas_ceder, key="sel_ceder")
                monto_ceder = st.number_input("Monto a Ceder ($)", min_value=0.0, max_value=float(presupuesto_lian_actual), step=50.0, value=0.0, key="inp_ceder")
                
                if st.button("Transferir Presupuesto", use_container_width=True):
                    if monto_ceder > 0:
                        ces_act = cfg_actual.get("cesion_lian", {})
                        ces_act[destinatario_ceder] = ces_act.get(destinatario_ceder, 0.0) + monto_ceder
                        cfg_actual["cesion_lian"] = ces_act
                        guardar_config(cfg_actual)
                        st.toast(f"Has transferido ${monto_ceder:,.2f} a {destinatario_ceder}", icon="🤝")
                        st.rerun()

        if cfg_actual.get("asignacion_comodin") or cfg_actual.get("cesion_lian"):
            if st.button("🔄 Restablecer Transferencias a Valores Originales", type="secondary", use_container_width=True):
                cfg_actual["asignacion_comodin"] = {}
                cfg_actual["cesion_lian"] = {}
                guardar_config(cfg_actual)
                st.toast("Transferencias restablecidas a presupuestos base.", icon="🔄")
                st.rerun()

    # Operadores catálogo para selectboxes
    todos_los_operadores = [""]
    for l_ops in OPERADORES_POR_SOLICITANTE.values():
        for o in l_ops:
            if o not in todos_los_operadores:
                todos_los_operadores.append(o)

    # 2. CARGA LUNES (FORMATO Y EDITOR FINAL)
    with tab_lunes:
        st.markdown("##### 🚗 Solicitud Oficial de Carga del LUNES")
        st.caption("Modifica libremente operadores y montos del **Lunes** antes de descargar:")
        
        df_lunes_view = df_actual.copy()
        df_lunes_view["Operador_Lunes"] = df_lunes_view["Operador_Lunes"].apply(limpiar_texto_operador)
        
        df_lunes_edit = st.data_editor(
            df_lunes_view,
            use_container_width=True,
            height=490,
            disabled=["row", "Solicitante", "Vehículo", "Placa", "Actividad", "Real_Lunes", "Operador_Jueves", "Importe_Jueves", "Real_Jueves"],
            column_config={
                "Solicitante": st.column_config.TextColumn("Área"),
                "Vehículo": st.column_config.TextColumn("Vehículo"),
                "Placa": st.column_config.TextColumn("Placa"),
                "Operador_Lunes": st.column_config.SelectboxColumn("Operador Lunes (Elegir)", options=todos_los_operadores, width="medium"),
                "Importe_Lunes": st.column_config.NumberColumn("Importe Lunes ($)", min_value=0.0, step=50.0, format="$%.2f"),
                "Actividad": st.column_config.TextColumn("Actividad", width="medium"),
                "row": None, "Real_Lunes": None, "Operador_Jueves": None, "Importe_Jueves": None, "Real_Jueves": None
            },
            hide_index=True,
            key="admin_editor_lunes"
        )
        
        if st.button("💾 Guardar Cambios del LUNES en Google Sheets", type="primary", use_container_width=True):
            with st.spinner("Guardando modificaciones del Lunes..."):
                exito = enviar_datos_sheets(df_lunes_edit, turno="lunes", f_elab=f_elab, f_prog=f_prog)
                if exito:
                    st.success("✅ ¡Cargas del Lunes guardadas y sincronizadas!")
                else:
                    st.error("Error al guardar en Google Sheets.")

        st.markdown("---")
        df_solo_lunes = df_lunes_edit[df_lunes_edit["Importe_Lunes"] > 0].copy()
        
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            excel_lunes = generar_excel_oficial_formato(df_lunes_edit, "Lunes", f_elab, f_prog)
            st.download_button(
                label="📥 Descargar Formato Excel LUNES (.xlsx)",
                data=excel_lunes,
                file_name=f"SOLICITUD_LUNES_{f_prog.strftime('%d%m%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_dl2:
            pdf_lunes = generar_pdf_oficial(df_solo_lunes, "Lunes", f_elab, f_prog)
            st.download_button(
                label="📄 Descargar Oficio PDF LUNES",
                data=pdf_lunes,
                file_name=f"OFICIO_LUNES_{f_prog.strftime('%d%m%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

    # 3. CARGA JUEVES (FORMATO Y EDITOR FINAL)
    with tab_jueves:
        st.markdown("##### 🚗 Solicitud Oficial de Carga del JUEVES")
        st.caption("Modifica libremente operadores y montos del **Jueves** antes de descargar:")
        
        df_jueves_view = df_actual.copy()
        df_jueves_view["Operador_Jueves"] = df_jueves_view["Operador_Jueves"].apply(limpiar_texto_operador)
        
        df_jueves_edit = st.data_editor(
            df_jueves_view,
            use_container_width=True,
            height=490,
            disabled=["row", "Solicitante", "Vehículo", "Placa", "Actividad", "Operador_Lunes", "Importe_Lunes", "Real_Lunes", "Real_Jueves"],
            column_config={
                "Solicitante": st.column_config.TextColumn("Área"),
                "Vehículo": st.column_config.TextColumn("Vehículo"),
                "Placa": st.column_config.TextColumn("Placa"),
                "Operador_Jueves": st.column_config.SelectboxColumn("Operador Jueves (Elegir)", options=todos_los_operadores, width="medium"),
                "Importe_Jueves": st.column_config.NumberColumn("Importe Jueves ($)", min_value=0.0, step=50.0, format="$%.2f"),
                "Actividad": st.column_config.TextColumn("Actividad", width="medium"),
                "row": None, "Operador_Lunes": None, "Importe_Lunes": None, "Real_Lunes": None, "Real_Jueves": None
            },
            hide_index=True,
            key="admin_editor_jueves"
        )
        
        if st.button("💾 Guardar Cambios del JUEVES en Google Sheets", type="primary", use_container_width=True):
            with st.spinner("Guardando modificaciones del Jueves..."):
                exito = enviar_datos_sheets(df_jueves_edit, turno="jueves", f_elab=f_elab, f_prog=f_prog)
                if exito:
                    st.success("✅ ¡Cargas del Jueves guardadas y sincronizadas!")
                else:
                    st.error("Error al guardar en Google Sheets.")

        st.markdown("---")
        df_solo_jueves = df_jueves_edit[df_jueves_edit["Importe_Jueves"] > 0].copy()
        
        col_dj1, col_dj2 = st.columns(2)
        with col_dj1:
            excel_jueves = generar_excel_oficial_formato(df_jueves_edit, "Jueves", f_elab, f_prog)
            st.download_button(
                label="📥 Descargar Formato Excel JUEVES (.xlsx)",
                data=excel_jueves,
                file_name=f"SOLICITUD_JUEVES_{f_prog.strftime('%d%m%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_dj2:
            pdf_jueves = generar_pdf_oficial(df_solo_jueves, "Jueves", f_elab, f_prog)
            st.download_button(
                label="📄 Descargar Oficio PDF JUEVES",
                data=pdf_jueves,
                file_name=f"OFICIO_JUEVES_{f_prog.strftime('%d%m%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

    # 4. MI CARGA (LIAN / FRANCISCO ALONZO)
    with tab_mi_carga:
        st.markdown("##### 🛵 Mi Carga (LIAN / Francisco Alonzo)")
        df_lian = df_actual[df_actual["Solicitante"] == "LIAN"].copy()
        presupuesto_lian_efectivo = presupuestos_actuales["LIAN"]
        
        if not df_lian.empty:
            row_lian = df_lian.iloc[0]
            val_imp_lunes = float(row_lian["Importe_Lunes"])
            val_imp_jueves = float(row_lian["Importe_Jueves"])
            
            with st.container(border=True):
                st.markdown(f"🛵 **{row_lian['Vehículo']}** &nbsp;|&nbsp; Placa: **`{row_lian['Placa']}`** &nbsp;|&nbsp; Presupuesto Autorizado: **${presupuesto_lian_efectivo:,.2f}**")
                
                c_ml1, c_ml2 = st.columns(2)
                with c_ml1:
                    st.markdown("🗓️ **Carga del Lunes**")
                    op_l_guardado = limpiar_texto_operador(row_lian["Operador_Lunes"])
                    op_l_sel = st.selectbox("Operador Lunes", ["", "FRANCISCO ALONZO"], index=1 if op_l_guardado == "FRANCISCO ALONZO" else 0, key="ml_op_l")
                    imp_l_sel = st.number_input("Monto Lunes ($)", value=val_imp_lunes, step=50.0, min_value=0.0, max_value=float(presupuesto_lian_efectivo), key="ml_imp_l")
                
                with c_ml2:
                    st.markdown("🗓️ **Carga del Jueves**")
                    disponible_jueves_lian = max(0.0, presupuesto_lian_efectivo - imp_l_sel)
                    op_j_guardado = limpiar_texto_operador(row_lian["Operador_Jueves"])
                    op_j_sel = st.selectbox("Operador Jueves", ["", "FRANCISCO ALONZO"], index=1 if op_j_guardado == "FRANCISCO ALONZO" else 0, key="ml_op_j")
                    imp_j_sel = st.number_input("Monto Jueves ($)", value=val_imp_jueves, step=50.0, min_value=0.0, max_value=float(disponible_jueves_lian), key="ml_imp_j")
                
                if st.button("💾 Guardar Mi Carga (Lunes y Jueves)", type="primary", use_container_width=True):
                    df_mi_carga = df_actual[df_actual["Solicitante"] == "LIAN"].copy()
                    df_mi_carga["Operador_Lunes"] = op_l_sel
                    df_mi_carga["Importe_Lunes"] = imp_l_sel
                    df_mi_carga["Operador_Jueves"] = op_j_sel
                    df_mi_carga["Importe_Jueves"] = imp_j_sel
                    
                    enviar_datos_sheets(df_mi_carga, turno="lunes", f_elab=f_elab, f_prog=f_prog)
                    enviar_datos_sheets(df_mi_carga, turno="jueves", f_elab=f_elab, f_prog=f_prog)
                    st.success("✅ Tu carga fue registrada correctamente.")

    # 5. AUDITORÍA Y COMPROBACIÓN REAL (LUNES Y JUEVES)
    with tab_auditoria:
        st.markdown("##### 🔍 Conciliación y Comprobación Real (Lunes y Jueves)")
        st.caption("Captura lo que realmente cargaron el Lunes y el Jueves para calcular el ahorro total:")
        
        df_audit_view = df_actual.copy()
        df_audit_view["Total_Solicitado_Semana"] = df_audit_view["Importe_Lunes"] + df_audit_view["Importe_Jueves"]
        
        df_audit_edit = st.data_editor(
            df_audit_view,
            use_container_width=True,
            height=490,
            disabled=["row", "Solicitante", "Vehículo", "Placa", "Actividad", "Importe_Lunes", "Importe_Jueves", "Total_Solicitado_Semana"],
            column_config={
                "Solicitante": st.column_config.TextColumn("Área"),
                "Vehículo": st.column_config.TextColumn("Vehículo"),
                "Placa": st.column_config.TextColumn("Placa"),
                "Importe_Lunes": st.column_config.NumberColumn("Sol. Lunes ($)", format="$%.2f"),
                "Real_Lunes": st.column_config.NumberColumn("Real Lunes ($)", min_value=0.0, step=50.0, format="$%.2f"),
                "Importe_Jueves": st.column_config.NumberColumn("Sol. Jueves ($)", format="$%.2f"),
                "Real_Jueves": st.column_config.NumberColumn("Real Jueves ($)", min_value=0.0, step=50.0, format="$%.2f"),
                "Total_Solicitado_Semana": st.column_config.NumberColumn("Total Solicitado ($)", format="$%.2f"),
                "row": None, "Actividad": None, "Operador_Lunes": None, "Operador_Jueves": None
            },
            hide_index=True,
            key="admin_editor_auditoria_completo"
        )
        
        total_sol_semana = df_audit_edit["Total_Solicitado_Semana"].sum()
        total_real_semana = df_audit_edit["Real_Lunes"].sum() + df_audit_edit["Real_Jueves"].sum()
        ahorro_semana = total_sol_semana - total_real_semana
        saldo_libre_global = PRESUPUESTO_GLOBAL - total_real_semana
        
        st.divider()
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Total Solicitado Semana", f"${total_sol_semana:,.2f}")
        r2.metric("Total Real Comprobado", f"${total_real_semana:,.2f}")
        r3.metric("Ahorro / Remanente", f"${ahorro_semana:,.2f}", delta_color="normal")
        r4.metric("Saldo Libre Total", f"${saldo_libre_global:,.2f}", delta_color="normal")
        
        st.write("")
        if st.button("💾 Guardar Comprobaciones Reales y Archivar en Histórico", type="primary", use_container_width=True):
            with st.spinner("Guardando en Google Sheets y archivando..."):
                enviar_datos_sheets(df_audit_edit, turno="real_lunes", f_elab=f_elab, f_prog=f_prog)
                
                detalles_txt_lista = []
                for _, r_c in df_audit_edit[(df_audit_edit['Real_Lunes'] > 0) | (df_audit_edit['Real_Jueves'] > 0)].iterrows():
                    tot_u = r_c['Real_Lunes'] + r_c['Real_Jueves']
                    detalles_txt_lista.append(f"{r_c['Vehículo']} - ${tot_u:,.2f}")
                detalles_txt = "; ".join(detalles_txt_lista) if detalles_txt_lista else "Sin cargas reales"
                
                folio_generado = f"SEMANA-{f_prog.strftime('%Y%m%d')}"
                historico_payload = {
                    "folio": folio_generado,
                    "fecha_elaboro": f_elab.strftime("%d/%m/%Y"),
                    "fecha_prog": f_prog.strftime("%d/%m/%Y"),
                    "fecha_registro": datetime.now(ZONA_HORARIA).strftime("%d/%m/%Y %I:%M %p"),
                    "total_solicitado": float(total_sol_semana),
                    "total_ejercido": float(total_real_semana),
                    "ahorro": float(ahorro_semana),
                    "detalle_unidades": detalles_txt
                }
                
                exito = enviar_datos_sheets(df_audit_edit, turno="real_jueves", f_elab=f_elab, f_prog=f_prog, historico_obj=historico_payload)
                if exito:
                    st.success(f"✅ ¡Auditoría sincronizada y Folio **{folio_generado}** archivado en 'historico'!")
                else:
                    st.error("Error al guardar en Google Sheets.")

    # 6. MODO PRUEBAS Y REESTABLECIMIENTO DE VIERNES
    with tab_mantenimiento:
        st.markdown("##### 🛠️ Mantenimiento y Cierre de Viernes")
        
        with st.container(border=True):
            st.subheader("🧹 Reestablecer Todo a $0.00 (Cada Viernes)")
            st.write(
                "Usa este botón los **viernes** tras guardar la auditoría semanal. Limpiará las cargas de Lunes y Jueves para arrancar la siguiente semana en ceros y con presupuestos completos."
            )
            
            if st.button("🗑️ Reestablecer Todo para la Próxima Semana", type="secondary", use_container_width=True):
                df_limpio = df_actual.copy()
                df_limpio["Operador_Lunes"] = ""
                df_limpio["Importe_Lunes"] = 0.0
                df_limpio["Real_Lunes"] = 0.0
                df_limpio["Operador_Jueves"] = ""
                df_limpio["Importe_Jueves"] = 0.0
                df_limpio["Real_Jueves"] = 0.0
                
                enviar_datos_sheets(df_limpio, turno="lunes", f_elab=f_elab, f_prog=f_prog)
                enviar_datos_sheets(df_limpio, turno="real_lunes", f_elab=f_elab, f_prog=f_prog)
                enviar_datos_sheets(df_limpio, turno="jueves", f_elab=f_elab, f_prog=f_prog)
                enviar_datos_sheets(df_limpio, turno="real_jueves", f_elab=f_elab, f_prog=f_prog)
                
                cfg_actual["dia_activo"] = "Lunes"
                guardar_config(cfg_actual)
                st.success("✅ ¡Semana reiniciada a $0.00 y lista para el próximo Lunes!")
                st.rerun()

        with st.container(border=True):
            st.subheader("🧪 Probar Vista Móvil de Solicitante")
            usuarios_para_test = [u for u in USUARIOS_PASSWORD.keys() if u != "LIAN"]
            solicitante_a_testear = st.selectbox("Selecciona al solicitante a simular:", usuarios_para_test)
            
            if st.button("👁️ Entrar a Modo Simulación", type="secondary", use_container_width=True):
                st.session_state.vista_simulada = solicitante_a_testear
                st.rerun()
